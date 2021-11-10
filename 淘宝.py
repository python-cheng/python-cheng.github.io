import requests
import re
import openpyxl
from multiprocessing.dummy import Pool

# 全局变量
headers = {"User-Agent":"Chrome/10","cookie": "cna=5ad8FxhGTBACAW9V2hajpuVd; miid=1868276696436572287; thw=cn; enc=GuGKSSX7YxZV%2B85AdF%2Bquc%2FtfC%2BWgLgGLk9d3uMDfJ8VriI%2BU%2B0QWy9Q5BPMEaLtJiYsnhMl33Eo%2BFLuux1w2Q%3D%3D; hng=CN%7Czh-CN%7CCNY%7C156; tracknick=%5Cu4EBA%5Cu751F%5Cu5C31%5Cu662F%5Cu4E00%5Cu573A%5Cu68A6zz; t=5d9256a8251b9ebfc43464c258bad689; cookie2=138cc40d94b840104c9a9b2003599a2f; _tb_token_=e3b77331ef937; v=0; _m_h5_tk=c04d1db924a84d7492130fedc3d67a47_1605098514753; _m_h5_tk_enc=f8986d6a3fbfcfccf68dc42178de1daa; _samesite_flag_=true; xlly_s=1; sgcookie=E1007vq1autySBv1sBNkfOstPD3uVCOh%2BufhnsMZZHqsP7%2BgrMrm1NESzCPEWT6B7tZBBOPJ3nzbREACsnydXJuzxQ%3D%3D; unb=2334961080; uc3=vt3=F8dCufOCQKLAlPHQlc0%3D&id2=UUtIEfLD8jP%2B1A%3D%3D&nk2=q5I3B1SRGWeUbxDRPZK%2B4g%3D%3D&lg2=V32FPkk%2Fw0dUvg%3D%3D; csg=4901c09b; lgc=%5Cu4EBA%5Cu751F%5Cu5C31%5Cu662F%5Cu4E00%5Cu573A%5Cu68A6zz; cookie17=UUtIEfLD8jP%2B1A%3D%3D; dnk=%5Cu4EBA%5Cu751F%5Cu5C31%5Cu662F%5Cu4E00%5Cu573A%5Cu68A6zz; skt=1985f8ec8efa3508; existShop=MTYwNTA5MDI3OQ%3D%3D; uc4=nk4=0%40qSdf8lyk3uOYUDyd5rqjNpMX4WSI%2FcZsVeGH&id4=0%40U2lyi2%2FqQalYEEx%2FZHR97yYoBDAE; _cc_=W5iHLLyFfA%3D%3D; _l_g_=Ug%3D%3D; sg=z00; _nk_=%5Cu4EBA%5Cu751F%5Cu5C31%5Cu662F%5Cu4E00%5Cu573A%5Cu68A6zz; cookie1=AiUNnGpW784T6G11vgmYRmh0WWT0Q9FTS5G7eX6%2By9c%3D; mt=ci=-1_1; uc1=cookie14=Uoe0aDy4VpQ7RQ%3D%3D&pas=0&cookie16=U%2BGCWk%2F74Mx5tgzv3dWpnhjPaQ%3D%3D&existShop=false&cookie15=URm48syIIVrSKA%3D%3D&cookie21=VFC%2FuZ9ainBZ; alitrackid=www.taobao.com; lastalitrackid=www.taobao.com; JSESSIONID=818915FD7FBB1F06A90696C2DB786C92; l=eBgph8R7QdR_JidSBOfwourza77tjIRA_uPzaNbMiOCP_R1p5uBhWZSiU189CnGVh6v6R3S4c1SwBeYBqIv4n5U68yvNI9Hmn; isg=BJubr1OdQIkks70jxP_j-AalKv8FcK9ydkcar43YfhqxbLtOFUBmwpjuAsxiywdq; tfstk=cA4RBPwi2tXoDaGY_uIDRht8L2fcZzS-KtDwJ1OyL74WeYTdiH4gWNYeofp-ksC.."}

# 爬取所有页面包含主页面的shop_url
page_url = []

for page in range(0,1960,20): 
    url = "https://shopsearch.taobao.com/search?q=迁西板栗&s=%d"%page
    # print(url)
    page_url.append(url)


name_list = []
_type_list = []
adderss_list = []
def requests_data(url):
    print('开始下载数据......')
    res = requests.get(url=url,headers=headers).text
    # 店名
    name = re.findall(r'"rawTitle":"(.*?)"',res)
    name_list.append(name)
    # 主营类型"mainAuction":"
    _type = re.findall(r'mainAuction":"(.*?)"',res)
    _type_list.append(_type)
    # 地址 "provcity":"河北 唐山"
    address = re.findall(r'"provcity":"(.*?)"',res)
    adderss_list.append(address)
    # print(address)
    return name_list,_type_list,adderss_list
    pass



pool = Pool(10)
a = pool.map(requests_data,page_url)[0]
pool.close()

# print(name[1])
name,_type,address = a
def write_inte_excel():
        n = []
        d = []
        t = []
        print("开始写入......")
        wb = openpyxl.Workbook()
        sh = wb.active
        sh.title ='迁西板栗'
        sh["A1"] = "店铺名称"
        sh["B1"] = "发货地"
        sh["C1"] = "主营类型"
        # print(sh.max_row)
        # print(name)
        row = 2
        for i in range(98):
            n += name[i]
            d += address[i]
            t += _type[i]
        for n1,a1,t1 in zip(n,d,t):
            sh.cell(row,1).value = n1
            sh.cell(row,2).value = a1
            sh.cell(row,3).value = t1
            row += 1
        print("写入完成！！！")
        wb.save('迁西板栗.xlsx')
write_inte_excel()








